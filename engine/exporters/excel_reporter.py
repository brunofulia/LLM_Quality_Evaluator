import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from engine.evaluation.domain_models import AuditResult
from engine.logger import get_logger

logger = get_logger(__name__)

class ExcelReporter:
    """
    Generates an Excel audit report with cell highlighting based on severity.
    """

    COLOR_MAP = {
        "PASS": "00FF00", # Green
        "MINOR_ISSUE": "FFFF00", # Yellow
        "MAJOR_ISSUE": "FFA500", # Orange
        "CRITICAL_AUDIT_FAILURE": "FF0000", # Red
    }

    @staticmethod
    def export(audit_result: AuditResult, output_dir: Path) -> Path:
        output_dir.mkdir(parents=True, exist_ok=True)
        file_path = output_dir / "audit_report.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Results"

        # Headers
        headers = ["Case ID", "Metric Name", "Input", "Actual Output", "Passed", "Severity", "Score", "Threshold", "Reason"]
        ws.append(headers)
        
        # Style Headers
        header_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data Rows
        for row_idx, case in enumerate(audit_result.case_results, start=2):
            row_data = [
                case.case_id,
                case.metric_name,
                case.input_text,
                case.actual_output,
                str(case.passed),
                case.severity,
                case.score,
                case.threshold,
                case.reason
            ]
            ws.append(row_data)

            # Apply severity color
            color_hex = ExcelReporter.COLOR_MAP.get(case.severity, "FFFFFF")
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            
            # Highlight the Severity cell specifically (Column F / 6)
            severity_cell = ws.cell(row=row_idx, column=6)
            severity_cell.fill = fill
            
            if case.severity == "CRITICAL_AUDIT_FAILURE":
                 severity_cell.font = Font(color="FFFFFF", bold=True)

        wb.save(file_path)
        logger.info(f"Excel report generated at {file_path}")
        return file_path
