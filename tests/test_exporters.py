import pytest
import os
from pathlib import Path
from unittest.mock import MagicMock
from engine.evaluation.domain_models import AuditResult, TestCaseResult
from engine.exporters.excel_reporter import ExcelReporter
from engine.exporters.html_reporter import HTMLReporter

@pytest.fixture
def dummy_audit_result():
    res1 = TestCaseResult(case_id="1", metric_name="Toxicity", input_text="<script>alert('xss')</script>", actual_output="safe", passed=True, severity="PASS", score=1.0, threshold=0.8, reason="ok")
    res2 = TestCaseResult(case_id="2", metric_name="Toxicity", input_text="in", actual_output="out", passed=False, severity="CRITICAL_AUDIT_FAILURE", score=0.1, threshold=0.8, reason="bad")
    return AuditResult(
        project_name="Test Project",
        provider="google",
        model="gemini-4o",
        case_results=[res1, res2],
        total_cases=2,
        passed_cases=1,
        failed_cases=1,
        critical_failures=1,
        execution_time_seconds=1.23
    )

def test_excel_reporter(dummy_audit_result, tmp_path):
    output_dir = tmp_path / "results"
    file_path = ExcelReporter.export(dummy_audit_result, output_dir)
    
    assert file_path.exists()
    assert file_path.name == "audit_report.xlsx"
    
    # We can use openpyxl to verify cell contents
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Headers should be at row 1
    assert ws.cell(row=1, column=1).value == "Case ID"
    
    # Data should be at row 2 and 3
    assert ws.cell(row=2, column=1).value == "1"
    assert ws.cell(row=2, column=2).value == "Toxicity"
    assert ws.cell(row=2, column=3).value == "<script>alert('xss')</script>"
    assert ws.cell(row=2, column=4).value == "safe"
    assert str(ws.cell(row=2, column=5).value) == "True"
    
    # Check severity color apply (Hex for CRITICAL is FF0000)
    # openpyxl returns '00FF0000' or similar due to alpha channel usually, so we just check it exists
    critical_cell = ws.cell(row=3, column=6)
    assert critical_cell.value == "CRITICAL_AUDIT_FAILURE"
    assert critical_cell.fill.start_color.rgb in ("FFFF0000", "00FF0000", "FF0000") # standard solid red

def test_html_reporter(dummy_audit_result, tmp_path):
    output_dir = tmp_path / "results"
    
    # Create a dummy template
    template_path = tmp_path / "report_template.html"
    template_path.write_text("Title: {{PROJECT_NAME}}, HTML: {{TABLE_ROWS}}, KO: {{KO_BANNER}}", encoding="utf-8")
    
    file_path = HTMLReporter.export(dummy_audit_result, output_dir, template_path)
    
    assert file_path.exists()
    
    content = file_path.read_text(encoding="utf-8")
    
    # Check injection
    assert "Title: Test Project" in content
    
    # Check XSS escaping
    assert "&lt;script&gt;alert(&#x27;xss&#x27;)&lt;/script&gt;" in content
    assert "<script>alert('xss')</script>" not in content
    
    # Check KO banner logic
    assert "CRITICAL KO DETECTED" in content

def test_html_reporter_missing_template(dummy_audit_result, tmp_path):
    output_dir = tmp_path / "results"
    template_path = tmp_path / "non_existent.html"
    
    with pytest.raises(FileNotFoundError):
        HTMLReporter.export(dummy_audit_result, output_dir, template_path)
